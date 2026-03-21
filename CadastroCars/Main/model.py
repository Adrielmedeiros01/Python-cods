import openpyxl as xl
import os

CAMINHO = "CadastroCars/Carros.xlsx"

class CarroModel:
    def __init__(self):
        if not os.path.exists("CadastroCars"):
            os.makedirs("CadastroCars")

        if not os.path.exists(CAMINHO):
            wb = xl.Workbook()
            ws = wb.active
            ws.title = "Planilha1"
            ws.append(["ID", "Marca", "Modelo"])
            wb.save(CAMINHO)

        self.workbook = xl.load_workbook(CAMINHO)
        self.sheet = self.workbook["Planilha1"]

    def listar(self):
        return list(self.sheet.iter_rows(min_row=2, values_only=True))

    def adicionar(self, marca, modelo):
        new_id = self.sheet.max_row
        self.sheet.append([new_id, marca.upper(), modelo.upper()])
        self.workbook.save(CAMINHO)

    def atualizar(self, id, marca, modelo):
        for row in self.sheet.iter_rows(min_row=2):
            if row[0].value == id:
                row[1].value = marca.upper()
                row[2].value = modelo.upper()
                self.workbook.save(CAMINHO)
                return True
        return False

    def deletar(self, id):
        for i, row in enumerate(self.sheet.iter_rows(min_row=2), start=2):
            if row[0].value == id:
                self.sheet.delete_rows(i)
                self.workbook.save(CAMINHO)
                return True
        return False
