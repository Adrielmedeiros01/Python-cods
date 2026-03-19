
import openpyxl as xl

workbook = xl.load_workbook("CadastroCars/Carros.xlsx")
carros = workbook["Planilha1"]

def visualizar():
    for i in carros.iter_rows(min_row=2):
        id = i[0].value
        marca = i[1].value
        modelo = i[2].value
        print(f"Id: {id} | Carro: {marca} | Modelo: {modelo}")

def adicionar():
    # gerar ID automático:
    id = carros.max_row
    marca = input("Marca: ").upper()
    modelo = input("Modelo: ").upper()
    carros.append([id, marca, modelo])
    workbook.save("CadastroCars/Carros.xlsx")
    print("Carro adicionado!")

def atualizar():
    visualizar()
    id_busca = int(input("\nDigite o ID para atualizar: "))
    for linha in carros.iter_rows(min_row=2):
        if linha[0].value == id_busca:
            linha[1].value = input("Nova marca: ")
            linha[2].value = input("Novo modelo: ")
            workbook.save("CadastroCars/Carros.xlsx")
            print("Atualizado com sucesso!")
            return

    print("ID não encontrado!")

def deletar():
    visualizar()
    id_busca = int(input("\nDigite o ID para deletar: "))
    for i, linha in enumerate(carros.iter_rows(min_row=2), start=2):
        if linha[0].value == id_busca:
            carros.delete_rows(i)
            workbook.save("CadastroCars/Carros.xlsx")
            print("Removido com sucesso!")
            return

    print("ID não encontrado!")

def menu():
    while True:
        print("\n--- MENU ---")
        print("1 - Adicionar")
        print("2 - Visualizar")
        print("3 - Atualizar")
        print("4 - Deletar")
        print("0 - Sair")

        op = input("Escolha: ")

        if op == "1":
            adicionar()
        elif op == "2":
            visualizar()
        elif op == "3":
            atualizar()
        elif op == "4":
            deletar()
        elif op == "0":
            print("Saindo...")
            break
        else:
            print("Opção inválida!")
menu()

# visualizar()
# adicionar()
# deletar()
# atualizar()
